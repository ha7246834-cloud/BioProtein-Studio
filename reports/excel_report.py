from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _style_worksheet(worksheet) -> None:
    """Apply simple professional formatting to an Excel worksheet."""

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    header_font = Font(
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        maximum_length = 0
        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            maximum_length = max(
                maximum_length,
                len(cell_value),
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            maximum_length + 3,
            45,
        )


def create_excel_report(
    properties_df: pd.DataFrame,
    amino_acid_df: pd.DataFrame,
    quality_df: pd.DataFrame,
    statistics_df: pd.DataFrame,
    correlation_df: pd.DataFrame,
    outliers_df: pd.DataFrame,
    interpretation_text: str,
) -> bytes:
    """
    Create a multi-sheet Excel workbook.

    Returns:
        Excel workbook as bytes.
    """

    output = BytesIO()

    interpretation_df = pd.DataFrame(
        {
            "Section": [
                "Physicochemical results interpretation",
            ],
            "Text": [
                interpretation_text,
            ],
        }
    )

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        properties_df.to_excel(
            writer,
            sheet_name="Physicochemical",
            index=False,
        )

        amino_acid_df.to_excel(
            writer,
            sheet_name="Amino_Acid_Composition",
            index=False,
        )

        quality_df.to_excel(
            writer,
            sheet_name="Quality_Control",
            index=False,
        )

        statistics_df.to_excel(
            writer,
            sheet_name="Summary_Statistics",
            index=False,
        )

        correlation_df.to_excel(
            writer,
            sheet_name="Correlation_Matrix",
            index=True,
        )

        if outliers_df is not None and not outliers_df.empty:
            outliers_df.to_excel(
                writer,
                sheet_name="Outliers",
                index=False,
            )
        else:
            pd.DataFrame(
                {
                    "Message": [
                        "No outliers were detected using the 1.5 × IQR method."
                    ]
                }
            ).to_excel(
                writer,
                sheet_name="Outliers",
                index=False,
            )

        interpretation_df.to_excel(
            writer,
            sheet_name="Interpretation",
            index=False,
        )

        for worksheet in writer.book.worksheets:
            _style_worksheet(worksheet)

        interpretation_sheet = writer.book[
            "Interpretation"
        ]

        interpretation_sheet.column_dimensions["A"].width = 35
        interpretation_sheet.column_dimensions["B"].width = 110
        interpretation_sheet.row_dimensions[2].height = 120

    output.seek(0)

    return output.getvalue()